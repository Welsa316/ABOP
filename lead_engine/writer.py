"""
writer.py — Output generation: enriched CSV, ranked CSV, JSON, text report,
             high-priority CSV, approved outreach queue, and Excel tracker.
"""

import csv
import logging
from pathlib import Path
from collections import Counter

from .utils import save_json, save_text
from . import config

logger = logging.getLogger("lead_engine")

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Columns for the output CSVs
OUTPUT_COLUMNS = [
    "contacted",
    "business_name",
    "primary_category",
    "city",
    "state",
    "phone",
    "website",
    "final_url",
    "instagram",
    "facebook",
    "tiktok",
    "email",
    "yelp",
    "contact_methods_found",
    "best_contact_channel",
    "instagram_confidence",
    "facebook_confidence",
    "tiktok_confidence",
    "email_confidence",
    "rating",
    "review_count",
    "lead_score",
    "opportunity_score",
    "reputation_score",
    "contactability_score",
    "digital_weakness_score",
    "priority_score",
    "score_summary",
    "recommended_pitch_label",
    "score_breakdown_text",
    "website_status",
    "detected_issues_text",
    "pitch_angle",
    "email_subject",
    "email_message",
    "contact_form_message",
    "dm_message",
    "follow_up_message",
    "call_script",
    "message_error",
    "status",
    "notes",
    "google_url",
]

# Excel column configuration: (field_key, header_label, width)
EXCEL_COLUMNS = [
    ("contacted",               "Contacted",            12),
    ("business_name",           "Business Name",        30),
    ("phone",                   "Phone Number",         16),
    ("email",                   "Email",                28),
    ("instagram",               "Instagram",            30),
    ("facebook",                "Facebook",             30),
    ("tiktok",                  "TikTok",               28),
    ("website",                 "Website",              30),
    ("primary_category",        "Category",             20),
    ("city",                    "City",                 16),
    ("state",                   "State",                 8),
    ("rating",                  "Rating",                8),
    ("review_count",            "Review Count",         12),
    ("lead_score",              "Lead Score",           11),
    ("opportunity_score",       "Opportunity",          12),
    ("reputation_score",        "Reputation",           11),
    ("contactability_score",    "Contactability",       14),
    ("digital_weakness_score",  "Digital Weakness",     16),
    ("recommended_pitch_label", "Recommended Pitch",    40),
    ("best_contact_channel",    "Best Channel",         14),
    ("dm_message",              "Generated DM Draft",   50),
    ("email_subject",           "Email Subject",        30),
    ("email_message",           "Generated Email Draft", 50),
    ("contact_form_message",    "Contact Form Draft",   50),
    ("follow_up_message",       "Follow-Up Draft",      50),
    ("call_script",             "Call Script",          50),
    ("status",                  "Status",               14),
    ("notes",                   "Notes",                30),
]


def _breakdown_text(breakdown: list[dict]) -> str:
    """Flatten score breakdown into a readable string."""
    if not breakdown:
        return ""
    parts = [f"{item['reason']} (+{item['points']})" if item['points'] > 0
             else f"{item['reason']} ({item['points']})"
             for item in breakdown]
    return "; ".join(parts)


def _issues_text(issues: list[str]) -> str:
    return ", ".join(issues) if issues else ""


def _biz_to_row(biz: dict) -> dict:
    """Convert a business dict into a flat row for CSV/Excel output."""
    return {
        "contacted":            biz.get("contacted", "No"),
        "business_name":        biz.get("business_name", ""),
        "primary_category":     biz.get("primary_category", ""),
        "city":                 biz.get("city", ""),
        "state":                biz.get("state", ""),
        "phone":                biz.get("phone", ""),
        "website":              biz.get("website", ""),
        "final_url":            biz.get("final_url", ""),
        "instagram":            biz.get("instagram", ""),
        "facebook":             biz.get("facebook", ""),
        "tiktok":               biz.get("tiktok", ""),
        "email":                biz.get("email", ""),
        "yelp":                 biz.get("yelp", ""),
        "contact_methods_found": biz.get("contact_methods_found", 0),
        "best_contact_channel": biz.get("best_contact_channel", ""),
        "instagram_confidence": biz.get("instagram_confidence", ""),
        "facebook_confidence":  biz.get("facebook_confidence", ""),
        "tiktok_confidence":    biz.get("tiktok_confidence", ""),
        "email_confidence":     biz.get("email_confidence", ""),
        "rating":               biz.get("rating", ""),
        "review_count":         biz.get("review_count", ""),
        "lead_score":           biz.get("lead_score", 0),
        "opportunity_score":    biz.get("opportunity_score", 0),
        "reputation_score":     biz.get("reputation_score", 0),
        "contactability_score": biz.get("contactability_score", 0),
        "digital_weakness_score": biz.get("digital_weakness_score", 0),
        "priority_score":       biz.get("priority_score", 0),
        "score_summary":        biz.get("score_summary", ""),
        "recommended_pitch_label": biz.get("recommended_pitch_label", ""),
        "score_breakdown_text": _breakdown_text(biz.get("score_breakdown", [])),
        "website_status":       biz.get("website_status", ""),
        "detected_issues_text": _issues_text(biz.get("detected_issues", [])),
        "pitch_angle":          biz.get("pitch_angle", ""),
        "email_subject":        biz.get("email_subject", ""),
        "email_message":        biz.get("email_message", ""),
        "contact_form_message": biz.get("contact_form_message", ""),
        "dm_message":           biz.get("dm_message", ""),
        "follow_up_message":    biz.get("follow_up_message", ""),
        "call_script":          biz.get("call_script", ""),
        "message_error":        biz.get("message_error", ""),
        "status":               biz.get("status", "new"),
        "notes":                biz.get("notes", ""),
        "google_url":           biz.get("google_url", ""),
    }


def _write_csv(rows: list[dict], path: Path) -> None:
    """Write a list of row dicts to a CSV file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)
    logger.info("Saved CSV → %s (%d rows)", path, len(rows))


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def _write_excel(businesses: list[dict], path: Path) -> None:
    """Write a formatted Excel workbook with Leads sheet and extras."""
    if not _HAS_OPENPYXL:
        logger.warning("openpyxl not installed — skipping Excel output. "
                       "Install with: pip install openpyxl")
        return

    wb = Workbook()

    # --- Main Leads sheet ---
    ws = wb.active
    ws.title = "Leads"
    _build_leads_sheet(ws, businesses)

    # --- High Priority sheet ---
    high_thresh = config.EXCEL_MEDIUM_SCORE_THRESHOLD
    high_priority = [b for b in businesses if b.get("lead_score", 0) >= high_thresh]
    if high_priority:
        ws_hp = wb.create_sheet("High Priority Leads")
        _build_leads_sheet(ws_hp, high_priority)

    # --- Summary sheet ---
    ws_summary = wb.create_sheet("Summary")
    _build_summary_sheet(ws_summary, businesses)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    logger.info("Saved Excel → %s (%d leads)", path, len(businesses))


def _build_leads_sheet(ws, businesses: list[dict]) -> None:
    """Populate a worksheet with lead data and formatting."""
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A4A8A", end_color="4A4A8A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # Write headers
    for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Write data rows
    for row_idx, biz in enumerate(businesses, 2):
        row = _biz_to_row(biz)
        for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            # Wrap text for message columns
            if "message" in key or "draft" in key or "script" in key or key in (
                "score_summary", "recommended_pitch_label", "follow_up_message",
                "dm_message", "email_message", "contact_form_message", "call_script"):
                cell.alignment = wrap_align

    total_rows = len(businesses) + 1  # +1 for header

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    if total_rows > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(EXCEL_COLUMNS))}{total_rows}"

    # Data validation for Contacted column (Yes/No dropdown)
    contacted_col = None
    for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
        if key == "contacted":
            contacted_col = col_idx
            break
    if contacted_col and total_rows > 1:
        dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv.error = "Please select Yes or No"
        dv.errorTitle = "Invalid Entry"
        col_letter = get_column_letter(contacted_col)
        dv.sqref = f"{col_letter}2:{col_letter}{total_rows}"
        ws.add_data_validation(dv)

    # Data validation for Status column
    status_col = None
    for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
        if key == "status":
            status_col = col_idx
            break
    if status_col and total_rows > 1:
        status_list = ",".join(config.CRM_STATUSES)
        dv_status = DataValidation(type="list", formula1=f'"{status_list}"',
                                   allow_blank=True)
        col_letter = get_column_letter(status_col)
        dv_status.sqref = f"{col_letter}2:{col_letter}{total_rows}"
        ws.add_data_validation(dv_status)

    # Conditional formatting for Lead Score column
    score_col = None
    for col_idx, (key, label, width) in enumerate(EXCEL_COLUMNS, 1):
        if key == "lead_score":
            score_col = col_idx
            break
    if score_col and total_rows > 1:
        col_letter = get_column_letter(score_col)
        score_range = f"{col_letter}2:{col_letter}{total_rows}"

        # Green for high scores
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        green_font = Font(color="006100", bold=True)
        ws.conditional_formatting.add(score_range, CellIsRule(
            operator="greaterThanOrEqual",
            formula=[str(config.EXCEL_HIGH_SCORE_THRESHOLD)],
            fill=green_fill, font=green_font,
        ))

        # Yellow for medium scores
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        yellow_font = Font(color="9C6500")
        ws.conditional_formatting.add(score_range, CellIsRule(
            operator="between",
            formula=[str(config.EXCEL_MEDIUM_SCORE_THRESHOLD),
                     str(config.EXCEL_HIGH_SCORE_THRESHOLD - 1)],
            fill=yellow_fill, font=yellow_font,
        ))

    # Conditional formatting for Contacted column
    if contacted_col and total_rows > 1:
        col_letter = get_column_letter(contacted_col)
        contacted_range = f"{col_letter}2:{col_letter}{total_rows}"
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        ws.conditional_formatting.add(contacted_range, CellIsRule(
            operator="equal",
            formula=['"Yes"'],
            fill=green_fill,
        ))


def _build_summary_sheet(ws, businesses: list[dict]) -> None:
    """Build a summary statistics sheet."""
    bold = Font(bold=True, size=12)
    header_font = Font(bold=True, size=14, color="4A4A8A")

    total = len(businesses)
    no_website = sum(1 for b in businesses if not b.get("website"))
    has_ig = sum(1 for b in businesses if b.get("instagram"))
    has_fb = sum(1 for b in businesses if b.get("facebook"))
    has_tt = sum(1 for b in businesses if b.get("tiktok"))
    has_em = sum(1 for b in businesses if b.get("email"))
    has_yelp = sum(1 for b in businesses if b.get("yelp"))
    has_any = sum(1 for b in businesses if b.get("contact_methods_found", 0) > 0)
    avg_score = sum(b.get("lead_score", 0) for b in businesses) / max(total, 1)
    high_priority = sum(1 for b in businesses
                        if b.get("lead_score", 0) >= config.EXCEL_MEDIUM_SCORE_THRESHOLD)

    # Count pitch angles
    angle_counter = Counter(b.get("pitch_angle", "") for b in businesses)

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20

    row = 1
    ws.cell(row=row, column=1, value="Lead Intelligence Summary").font = header_font
    row += 2

    stats = [
        ("Total Leads Processed", total),
        ("No Website", no_website),
        ("Instagram Found", has_ig),
        ("Facebook Found", has_fb),
        ("TikTok Found", has_tt),
        ("Email Found", has_em),
        ("Yelp Found", has_yelp),
        ("Any Contact Method", has_any),
        ("Average Lead Score", f"{avg_score:.1f}"),
        ("High Priority Leads", high_priority),
    ]
    for label, value in stats:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top Pitch Angles").font = header_font
    row += 1
    for angle, count in angle_counter.most_common(10):
        label = config.PITCH_ANGLE_LABELS.get(angle, angle)
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=count)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Top 10 Leads").font = header_font
    row += 1
    ws.cell(row=row, column=1, value="Business Name").font = bold
    ws.cell(row=row, column=2, value="Score").font = bold
    ws.cell(row=row, column=3, value="Pitch").font = bold
    ws.column_dimensions["C"].width = 45
    row += 1
    for b in businesses[:10]:
        ws.cell(row=row, column=1, value=b.get("business_name", ""))
        ws.cell(row=row, column=2, value=b.get("lead_score", 0))
        ws.cell(row=row, column=3, value=b.get("recommended_pitch_label", ""))
        row += 1


# ---------------------------------------------------------------------------
# Report
# ---------------------------------------------------------------------------

def _build_report(businesses: list[dict]) -> str:
    """Build a plain-text summary report."""
    total = len(businesses)
    no_website = sum(1 for b in businesses if not b.get("website"))
    unreachable = sum(1 for b in businesses
                      if b.get("website_status") == "unreachable")
    social_only = sum(1 for b in businesses
                      if b.get("website_status") == "social_only")

    # Count issue frequencies
    issue_counter: Counter = Counter()
    for b in businesses:
        for issue in b.get("detected_issues", []):
            issue_counter[issue] += 1

    # Count pitch angles
    angle_counter: Counter = Counter()
    for b in businesses:
        angle = b.get("pitch_angle", "")
        if angle:
            angle_counter[angle] += 1

    top20 = businesses[:20]

    # Contact discovery stats
    has_ig = sum(1 for b in businesses if b.get("instagram"))
    has_fb = sum(1 for b in businesses if b.get("facebook"))
    has_tt = sum(1 for b in businesses if b.get("tiktok"))
    has_em = sum(1 for b in businesses if b.get("email"))
    has_yelp = sum(1 for b in businesses if b.get("yelp"))
    has_any_contact = sum(1 for b in businesses if b.get("contact_methods_found", 0) > 0)

    avg_score = sum(b.get("lead_score", 0) for b in businesses) / max(total, 1)
    high_priority = sum(1 for b in businesses if b.get("lead_score", 0) >= 30)

    lines = [
        "=" * 60,
        "  LEAD INTELLIGENCE SUMMARY REPORT",
        "=" * 60,
        "",
        f"Total businesses processed:   {total}",
        f"Businesses with NO website:   {no_website}",
        f"Websites unreachable/broken:  {unreachable}",
        f"Social-media-only profiles:   {social_only}",
        f"Average lead score:           {avg_score:.1f}",
        f"High priority leads (30+):    {high_priority}",
        f"Leads ready for outreach:     {has_any_contact}",
        "",
        "  CONTACT DISCOVERY",
        f"  Instagram found:            {has_ig}",
        f"  Facebook found:             {has_fb}",
        f"  TikTok found:               {has_tt}",
        f"  Email found:                {has_em}",
        f"  Yelp found:                 {has_yelp}",
        f"  Any contact method:         {has_any_contact} / {total}",
        "",
        "-" * 60,
        "  TOP PITCH ANGLES",
        "-" * 60,
    ]
    for angle, count in angle_counter.most_common(10):
        label = config.PITCH_ANGLE_LABELS.get(angle, angle)
        lines.append(f"  {label:45s}  {count:>4d}")

    if issue_counter:
        lines.extend([
            "",
            "-" * 60,
            "  TOP SCORING REASONS",
            "-" * 60,
        ])
        for issue, count in issue_counter.most_common(15):
            lines.append(f"  {issue:30s}  {count:>4d} businesses")

    lines.extend([
        "",
        "-" * 60,
        "  TOP 20 LEADS",
        "-" * 60,
    ])
    for i, b in enumerate(top20, 1):
        name = b.get("business_name", "?")
        score = b.get("lead_score", 0)
        angle = b.get("recommended_pitch_label", b.get("pitch_angle", ""))
        city = b.get("city", "")
        channel = b.get("best_contact_channel", "")
        lines.append(f"  {i:>2}. [{score:>3} pts]  {name}  ({city})")
        if angle:
            lines.append(f"      Angle: {angle}")
        if channel:
            lines.append(f"      Best contact: {channel}")

    lines.extend(["", "=" * 60, "  END OF REPORT", "=" * 60, ""])
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Main output function
# ---------------------------------------------------------------------------

def write_outputs(businesses: list[dict], output_dir: str | Path) -> dict:
    """
    Write all output files.

    Returns a dict of {file_type: Path} for the files created.
    """
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    rows = [_biz_to_row(b) for b in businesses]

    files: dict[str, Path] = {}

    # 1. Enriched master CSV
    enriched_path = out / "enriched_master.csv"
    _write_csv(rows, enriched_path)
    files["enriched_csv"] = enriched_path

    # 2. Ranked leads CSV
    ranked_path = out / "ranked_leads.csv"
    _write_csv(rows, ranked_path)
    files["ranked_csv"] = ranked_path

    # 3. High-priority CSV (score >= threshold)
    high_priority = [r for r in rows
                     if int(r.get("lead_score", 0)) >= config.EXCEL_MEDIUM_SCORE_THRESHOLD]
    if high_priority:
        hp_path = out / "high_priority_leads.csv"
        _write_csv(high_priority, hp_path)
        files["high_priority_csv"] = hp_path

    # 4. Approved outreach queue (status == approved or contacted)
    approved = [r for r in rows if r.get("status") in ("approved", "contacted")]
    if approved:
        approved_path = out / "approved_outreach_queue.csv"
        _write_csv(approved, approved_path)
        files["approved_csv"] = approved_path

    # 5. JSON with full structured data
    json_data = []
    for b in businesses:
        entry = {k: v for k, v in b.items() if k != "_raw"}
        json_data.append(entry)
    json_path = out / "lead_data.json"
    save_json(json_data, json_path)
    files["json"] = json_path

    # 6. Text summary report
    report = _build_report(businesses)
    report_path = out / "summary_report.txt"
    save_text(report, report_path)
    files["report"] = report_path

    # 7. Excel tracker
    excel_path = out / config.EXCEL_FILENAME
    _write_excel(businesses, excel_path)
    files["excel"] = excel_path

    # Print report to console too
    print("\n" + report)

    return files
